"""
Импорт сотрудников и объектов из Excel-файлов в базу данных.
"""
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from app.core.database import db_manager

logger = logging.getLogger(__name__)


def _s_val(val) -> str:
    """Приводит значение ячейки Excel к строке."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    return str(val).strip()


def import_employees_from_excel(path: Path) -> int:
    """
    Импортирует сотрудников из Excel (штатное расписание).
    Возвращает количество обработанных записей.
    """
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [_s_val(c).lower() for c in header_row]

    def find_col(substr: str) -> Optional[int]:
        substr_lower = substr.lower()
        for i, h in enumerate(headers):
            if substr_lower in h:
                return i
        return None

    idx_tbn = find_col("табельный номер")
    idx_fio = find_col("сотрудник")
    idx_pos = find_col("должность")
    idx_dep = find_col("подразделение")
    idx_dismissal = find_col("увольн")

    if idx_fio is None or idx_tbn is None:
        raise RuntimeError(
            "Не найдены обязательные колонки "
            "'Табельный номер' и/или 'Сотрудник'"
        )

    processed = 0

    with db_manager.connection() as conn:
        with conn.cursor() as cur:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                fio = _s_val(row[idx_fio]) if idx_fio < len(row) else ""
                tbn = _s_val(row[idx_tbn]) if idx_tbn < len(row) else ""
                pos = (
                    _s_val(row[idx_pos])
                    if idx_pos is not None and idx_pos < len(row)
                    else ""
                )
                dep_name = (
                    _s_val(row[idx_dep])
                    if idx_dep is not None and idx_dep < len(row)
                    else ""
                )
                dismissal_raw = (
                    row[idx_dismissal]
                    if idx_dismissal is not None and idx_dismissal < len(row)
                    else None
                )

                if not fio and not tbn:
                    continue

                is_fired = bool(dismissal_raw and _s_val(dismissal_raw))

                department_id = None
                if dep_name:
                    cur.execute(
                        "SELECT id FROM departments WHERE name = %s",
                        (dep_name,),
                    )
                    r = cur.fetchone()
                    if r:
                        department_id = r[0]
                    else:
                        cur.execute(
                            "INSERT INTO departments (name) VALUES (%s) RETURNING id",
                            (dep_name,),
                        )
                        department_id = cur.fetchone()[0]

                if tbn:
                    cur.execute("SELECT id FROM employees WHERE tbn = %s", (tbn,))
                else:
                    cur.execute("SELECT id FROM employees WHERE fio = %s", (fio,))
                r = cur.fetchone()

                if r:
                    cur.execute(
                        """
                        UPDATE employees
                        SET fio = %s, tbn = %s, position = %s,
                            department_id = %s, is_fired = %s
                        WHERE id = %s
                        """,
                        (fio or None, tbn or None, pos or None,
                         department_id, is_fired, r[0]),
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO employees
                            (fio, tbn, position, department_id, is_fired)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (fio or None, tbn or None, pos or None,
                         department_id, is_fired),
                    )
                processed += 1

        conn.commit()

    wb.close()
    logger.info("Импорт сотрудников завершён: %d записей", processed)
    return processed


def import_objects_from_excel(path: Path) -> int:
    """
    Импортирует/обновляет объекты из Excel-справочника.
    Возвращает количество обработанных записей.
    """
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row_idx = None
    headers = []

    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1
    ):
        if row and any(
            "id (код) номер объекта" in _s_val(c).lower() for c in row
        ):
            header_row_idx = i
            headers = [_s_val(c).lower() for c in row]
            break

    if header_row_idx is None:
        raise RuntimeError(
            "Не найдена строка заголовка с колонкой 'ID (код) номер объекта'"
        )

    def find_col(substr: str) -> Optional[int]:
        substr_lower = substr.lower()
        for i, h in enumerate(headers):
            if substr_lower in h:
                return i
        return None

    idx_excel_id = find_col("id (код) номер объекта")
    idx_year = find_col("год")
    idx_program = find_col("наименование программы")
    idx_customer = find_col("наименование заказчика")
    idx_addr = find_col("адрес")
    idx_contract_no = find_col("№ договора")
    idx_contract_date = find_col("дата договора")
    idx_short_name = find_col("сокращенное наименование объекта")
    idx_department_name = find_col("подразделение")
    idx_contract_type = find_col("тип договора")

    if idx_excel_id is None or idx_addr is None:
        raise RuntimeError(
            "Не найдены обязательные колонки "
            "'ID (код) номер объекта' и/или 'Адрес объекта'"
        )

    processed = 0

    with db_manager.connection() as conn:
        with conn.cursor() as cur:
            for row in ws.iter_rows(
                min_row=header_row_idx + 1, values_only=True
            ):
                if not row:
                    continue

                excel_id = (
                    _s_val(row[idx_excel_id])
                    if idx_excel_id < len(row) else ""
                )
                if not excel_id:
                    continue

                year = (
                    _s_val(row[idx_year])
                    if idx_year is not None and idx_year < len(row) else ""
                )
                program = (
                    _s_val(row[idx_program])
                    if idx_program is not None and idx_program < len(row) else ""
                )
                customer_name = (
                    _s_val(row[idx_customer])
                    if idx_customer is not None and idx_customer < len(row) else ""
                )
                address = (
                    _s_val(row[idx_addr])
                    if idx_addr < len(row) else ""
                )
                contract_number = (
                    _s_val(row[idx_contract_no])
                    if idx_contract_no is not None and idx_contract_no < len(row)
                    else ""
                )

                contract_date_raw = (
                    row[idx_contract_date]
                    if idx_contract_date is not None
                    and idx_contract_date < len(row)
                    else None
                )
                contract_date_val: Optional[date] = None
                if isinstance(contract_date_raw, datetime):
                    contract_date_val = contract_date_raw.date()

                short_name = (
                    _s_val(row[idx_short_name])
                    if idx_short_name is not None
                    and idx_short_name < len(row) else ""
                )
                executor_department = (
                    _s_val(row[idx_department_name])
                    if idx_department_name is not None
                    and idx_department_name < len(row) else ""
                )
                contract_type = (
                    _s_val(row[idx_contract_type])
                    if idx_contract_type is not None
                    and idx_contract_type < len(row) else ""
                )

                cur.execute(
                    "SELECT id FROM objects WHERE excel_id = %s",
                    (excel_id,),
                )
                existing = cur.fetchone()

                if existing:
                    cur.execute(
                        """
                        UPDATE objects SET
                            address = %s, year = %s, program_name = %s,
                            customer_name = %s, contract_number = %s,
                            contract_date = %s, short_name = %s,
                            executor_department = %s, contract_type = %s,
                            updated_at = NOW()
                        WHERE id = %s
                        """,
                        (
                            address, year or None, program or None,
                            customer_name or None, contract_number or None,
                            contract_date_val, short_name or None,
                            executor_department or None,
                            contract_type or None, existing[0],
                        ),
                    )
                else:
                    cur.execute(
                        """
                        INSERT INTO objects
                            (excel_id, address, year, program_name,
                             customer_name, contract_number, contract_date,
                             short_name, executor_department,
                             contract_type, status)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """,
                        (
                            excel_id, address, year or None,
                            program or None, customer_name or None,
                            contract_number or None, contract_date_val,
                            short_name or None,
                            executor_department or None,
                            contract_type or None, "Новый",
                        ),
                    )
                processed += 1

        conn.commit()

    wb.close()
    logger.info("Импорт объектов завершён: %d записей", processed)
    return processed
